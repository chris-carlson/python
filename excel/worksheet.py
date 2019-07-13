from typing import List

from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet


class Worksheet:

    @staticmethod
    def _row_is_blank(row: List[str]) -> bool:
        return len([cell for cell in row if cell is not None]) == 0

    def __init__(self, openpyxl_worksheet: OpenpyxlWorksheet) -> None:
        self._openpyxl_worksheet: OpenpyxlWorksheet = openpyxl_worksheet
        self._rows: List[List[str]] = []

    def __len__(self) -> int:
        return len(self._rows)

    def __iter__(self) -> List[str]:
        for row in self._rows:
            yield row

    def __getitem__(self, index: int) -> List[str]:
        return self._rows[index]

    @property
    def title(self):
        return self._openpyxl_worksheet.title

    @title.setter
    def title(self, title):
        self._openpyxl_worksheet.title = title

    def get_column_by_header(self, header):
        first_row = self._rows[0]
        for column_index in range(0, len(first_row)):
            if first_row[column_index] == header:
                return self.get_column_by_index(column_index)[1:]
        raise ValueError('Could not find column with header \'' + header + '\'')

    def get_column_by_index(self, index: int) -> List[str]:
        if index < 0:
            raise IndexError('Index \'' + str(index) + '\' must be positive')
        column: List[str] = []
        for row in self._rows:
            if index >= len(row):
                raise IndexError('Index \'' + str(index) + '\' must be less than the number of columns')
            column.append(row[index])
        return column

    def add_row(self, values: List[str]) -> None:
        self._rows.append(values)

    def insert_row(self, index: int, values: List[str]) -> None:
        if index < 0:
            raise IndexError('Index \'' + str(index) + '\' must be positive')
        if index > len(self._rows):
            raise IndexError('Index \'' + str(index) + '\' must be less than or equal to the number of rows')
        self._rows.insert(index, values)

    def replace_row(self, index: int, values: List[str]) -> None:
        if index < 0:
            raise IndexError('Index \'' + str(index) + '\' must be positive')
        if index >= len(self._rows):
            raise IndexError('Index \'' + str(index) + '\' must be less than the number of rows')
        self._rows[index] = values

    def delete_row(self, index: int) -> None:
        if index < 0:
            raise IndexError('Index \'' + str(index) + '\' must be positive')
        if index >= len(self._rows):
            raise IndexError('Index \'' + str(index) + '\' must be less than the number of rows')
        self._rows.pop(index)

    def delete_rows(self, start_index: int = None, end_index: int = None) -> None:
        if start_index is None:
            start_index = 0
        if end_index is None:
            end_index = len(self._rows)
        if start_index < 0:
            raise IndexError('Start index \'' + str(start_index) + '\' must be positive')
        if end_index < 0:
            raise IndexError('End index \'' + str(end_index) + '\' must be positive')
        if start_index >= len(self._rows):
            raise IndexError('Start index \'' + str(start_index) + '\' must be less than the number of rows')
        if end_index >= len(self._rows):
            raise IndexError('End index \'' + str(end_index) + '\' must be less than the number of rows')
        if start_index > end_index:
            raise IndexError(
                'Start index \'' + str(start_index) + '\' must be less than the end index \'' + str(end_index) + '\'')
        for index in range(start_index, end_index):
            self.delete_row(start_index)

    def clear_row(self, index: int) -> None:
        if index < 0:
            raise IndexError('Index \'' + str(index) + '\' must be positive')
        if index >= len(self._rows):
            raise IndexError('Index \'' + str(index) + '\' must be less than the number of rows')
        self._rows[index] = []

    def clear_rows(self, start_index: int = None, end_index: int = None) -> None:
        if start_index is None:
            start_index = 0
        if end_index is None:
            end_index = len(self._rows)
        if start_index < 0:
            raise IndexError('Start index \'' + str(start_index) + '\' must be positive')
        if end_index < 0:
            raise IndexError('End index \'' + str(end_index) + '\' must be positive')
        if start_index >= len(self._rows):
            raise IndexError('Start index \'' + str(start_index) + '\' must be less than the number of rows')
        if end_index >= len(self._rows):
            raise IndexError('End index \'' + str(end_index) + '\' must be less than the number of rows')
        if start_index > end_index:
            raise IndexError(
                'Start index \'' + str(start_index) + '\' must be less than the end index \'' + str(end_index) + '\'')
        for index in range(start_index, end_index):
            self.clear_row(index)

    def load_data(self) -> None:
        for openpyxl_row in self._openpyxl_worksheet.rows:
            row: List[str] = []
            for openpyxl_cell in openpyxl_row:
                row.append(openpyxl_cell.value)
            self._rows.append(row)
        while len(self._rows) > 0 and self._row_is_blank(self._rows[-1]):
            self._rows.pop(-1)

    def save(self) -> None:
        for row_index in range(0, self._openpyxl_worksheet.max_row):
            openpyxl_row: List[str] = self._openpyxl_worksheet[row_index + 1]
            for column_index in range(0, len(openpyxl_row)):
                self._openpyxl_worksheet.cell(row=row_index + 1, column=column_index + 1).value = None
        for row_index in range(0, len(self._rows)):
            row: List[str] = self._rows[row_index]
            for column_index in range(0, len(row)):
                self._openpyxl_worksheet.cell(row=row_index + 1, column=column_index + 1, value=row[column_index])
