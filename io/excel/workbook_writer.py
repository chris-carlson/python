from typing import List

from cac.io.excel.cell import Cell
from cac.io.excel.row import Row
from cac.io.excel.worksheet import Worksheet
from openpyxl import Workbook as OpenpyxlWorkbook
from openpyxl.cell import Cell as OpenpyxlCell
from openpyxl.styles import Font, PatternFill
from openpyxl.worksheet.worksheet import Worksheet as OpenpyxlWorksheet

class WorkbookWriter:

    def __init__(self, file_name: str) -> None:
        self._file_name: str = file_name
        self._openpyxl_workbook: OpenpyxlWorkbook = OpenpyxlWorkbook()
        self._openpyxl_workbook.remove(self._openpyxl_workbook.worksheets[0])

    def write_worksheets(self, worksheets: List[Worksheet]) -> None:
        for worksheet in worksheets:
            self.write_worksheet(worksheet)

    # noinspection PyDunderSlots,PyUnresolvedReferences
    def write_worksheet(self, worksheet: Worksheet) -> None:
        openpyxl_worksheet: OpenpyxlWorksheet = self._openpyxl_workbook.create_sheet(worksheet.title)
        for row_index in range(0, len(worksheet)):
            row: Row = worksheet[row_index]
            for column_index in range(0, len(row)):
                cell: Cell = row[column_index]
                openpyxl_cell: OpenpyxlCell = openpyxl_worksheet.cell(row=row_index + 1, column=column_index + 1,
                    value=cell.value)
                if cell.bold:
                    openpyxl_cell.font = Font(bold=True)
                if cell.color:
                    openpyxl_cell.fill = PatternFill(start_color=cell.color, end_color=cell.color, fill_type='solid')
                if cell.number_format:
                    openpyxl_cell.number_format = cell.number_format
        self._openpyxl_workbook.save(self._file_name)

    def close(self) -> None:
        self._openpyxl_workbook.close()
