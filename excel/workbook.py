from typing import List

from cac.excel.worksheet import Worksheet
from openpyxl import Workbook as OpenpyxlWorkbook, load_workbook


class Workbook:

    def __init__(self, file_name: str = None) -> None:
        self._file_name: str = file_name
        self._worksheets: List[Worksheet] = []
        self._openpyxl_workbook: OpenpyxlWorkbook = OpenpyxlWorkbook()

    def __len__(self) -> int:
        return len(self._worksheets)

    def __iter__(self) -> Worksheet:
        for worksheet in self._worksheets:
            yield worksheet

    def __getitem__(self, index: int) -> Worksheet:
        if not type(index) == int:
            raise TypeError('Index \'' + str(index) + '\' must be of integer type')
        if index < 0:
            raise IndexError('Index \'' + str(index) + '\' must be positive')
        if index >= len(self._worksheets):
            raise IndexError('Index \'' + str(index) + '\' must be less than the number of worksheets')
        return self._worksheets[index]

    def load_data(self) -> None:
        if self._file_name is None:
            raise ValueError('Could not load data for workbook. Please provide file name.')
        self._openpyxl_workbook = load_workbook(filename=self._file_name)
        for openpyxl_worksheet in self._openpyxl_workbook:
            worksheet: Worksheet = Worksheet(openpyxl_worksheet)
            worksheet.load_data()
            self._worksheets.append(worksheet)

    def get_worksheet(self, title: str) -> Worksheet:
        worksheets: List[Worksheet] = [worksheet for worksheet in self._worksheets if worksheet.title == title]
        if len(worksheets) == 0:
            raise ValueError('Could not find worksheet with title \'' + title + '\'')
        elif len(worksheets) > 1:
            raise ValueError('Found multiple worksheets with title \'' + title + '\'')
        return worksheets[0]

    def create_worksheet(self, title: str, position: int = None) -> Worksheet:
        if position is None:
            position: int = len(self._worksheets)
        if position < 0:
            raise IndexError('Position \'' + str(position) + '\' must be positive')
        if position > len(self._worksheets):
            raise IndexError(
                'Position \'' + str(position) + '\' must be less than or equal to the number of worksheets')
        openpyxl_worksheet: OpenpyxlWorkbook = self._openpyxl_workbook.create_sheet(title, position)
        self._worksheets.insert(position, Worksheet(openpyxl_worksheet))
        return self._worksheets[position]

    def save(self) -> None:
        if self._file_name is None:
            raise ValueError('Could not save workbook. Please provide file name.')
        self._save_worksheets()
        self._openpyxl_workbook.save(self._file_name)

    def save_as(self, file_name: str) -> None:
        self._save_worksheets()
        self._openpyxl_workbook.save(file_name)

    def _save_worksheets(self) -> None:
        for worksheet in self._worksheets:
            worksheet.save()
